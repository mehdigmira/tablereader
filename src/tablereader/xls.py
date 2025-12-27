from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _get_max_row_with_data(ws: Worksheet):
    max_row = 0
    for i, row in enumerate(ws.rows):
        if any(cell.value is not None for cell in row):
            max_row = i
    return max_row


def _get_row_str(idx, row):
    row_str = " | ".join(
        str(cell.value) if cell.value is not None else "" for cell in row
    )
    return f"Row {idx}: {row_str}"


def get_sheet_content(file_path: str | Path, sheet_name: str):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    row_count = _get_max_row_with_data(ws) + 1

    content = []
    content.append(f"Excel file: {file_path}")
    content.append(f"Sheet: {sheet_name}")
    content.append(f"Total rows: {row_count}")

    if row_count > 200:
        # only return preview
        content.append("Sheet preview (only first and last 20 rows):")
        content.append("--------------------------------")
        # keep only 20 first rows and 20 last rows
        for idx, row in enumerate(ws.iter_rows(max_row=20)):
            content.append(_get_row_str(idx, row))

        content.append("\n...\n")

        for idx, row in enumerate(
            ws.iter_rows(min_row=row_count - 20, max_row=row_count)
        ):
            content.append(_get_row_str(row_count - 20 + idx - 1, row))

    else:
        content.append("Sheet full content:")
        content.append("--------------------------------")
        for i, row in enumerate(ws.iter_rows(max_row=row_count)):
            content.append(_get_row_str(i, row))

    wb.close()

    return "\n".join(content)


def table_iterator(
    data_start_row: int,
    data_end_row: int,
    start_col: int,
    end_col: int,
    skip_rows: list,
    *,
    file_path: str | Path,
    sheet_name: str,
):
    """
    Deterministically extract table to CSV based on LLM's analysis.

    Uses openpyxl directly to stream data without loading entire file into memory.
    Searches for the header row by matching column names.
    """

    # Open Excel file with openpyxl (read-only mode)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # Convert to set for faster lookup
    skip_rows_set = set(skip_rows) if skip_rows else set()

    # Iterate through rows
    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=data_start_row + 1,
            max_row=data_end_row + 1,
            min_col=start_col + 1,
            max_col=end_col + 1,
        )
    ):
        normalized_row = [
            str(cell.value).strip() if cell.value is not None else "" for cell in row
        ]
        # Skip specified rows
        if row_idx in skip_rows_set:
            continue
        yield normalized_row

    wb.close()
